from flask import Flask, render_template, request, redirect, flash, url_for
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from config import Config
from extensions import db, login_manager

from models.user_settings import UserSettings

from models.user import User
from models.expense import Expense
from models.budget import Budget
from models.income import Income
from models.goal import Goal
from models.notification import Notification
from services.notification_service import (
    generate_notifications,
    generate_budget_notifications,
    generate_goal_notifications,
    generate_income_notifications,
    generate_expense_notifications,
    generate_investment_notifications,
    generate_savings_notifications,
    generate_financial_health_notifications,
    generate_spending_insights,
)


from flask import send_file
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO

from datetime import datetime, timedelta
from sqlalchemy import extract

import requests

import os
from werkzeug.utils import secure_filename


from models.investment import Investment
from models.investment_transaction import InvestmentTransaction
from market_service import (
    get_stock_price,
    refresh_investment_prices,
    get_market_overview
)

from market_service import (
    get_stock_price,
    refresh_investment_prices,
    get_market_overview,
    get_historical_price
)

from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer

from utils.helpers import (
    time_ago,
    convert_amount,
    convert_to_base_currency,
    convert_list,
    get_currency_symbol
)

app = Flask(__name__)
app.config["SECRET_KEY"] = "finsight-secret-key-2026"
app.config["MAIL_SERVER"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True

app.config["MAIL_USERNAME"] = "finshight@gmail.com"
app.config["MAIL_PASSWORD"] = "ljmcmastczlcqzxb"

app.config["MAIL_DEFAULT_SENDER"] = (
    "FinSight",
    app.config["MAIL_USERNAME"]
)

mail = Mail(app)

serializer = URLSafeTimedSerializer(
    app.config["SECRET_KEY"]
)
app.config.from_object(Config)

db.init_app(app)
login_manager.init_app(app)

login_manager.login_view = "login"




@login_manager.user_loader
def load_user(user_id):
    return db.session.get(
        User,
        int(user_id)
    )


# -------------------------
# HOME
# -------------------------

@app.route("/")
def home():
    return redirect("/register")


# -------------------------
# DASHBOARD
# -------------------------

@app.route("/dashboard")
@login_required
def dashboard():

    from datetime import date

    today = date.today()

    total_expense = db.session.query(
        db.func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.user_id,
        db.extract("month", Expense.expense_date) == today.month,
        db.extract("year", Expense.expense_date) == today.year
    ).scalar()

    if total_expense is None:
        total_expense = 0

    monthly_income = db.session.query(
        db.func.sum(Income.amount)
    ).filter(
        Income.user_id == current_user.user_id,
        db.extract("month", Income.income_date) == today.month,
        db.extract("year", Income.income_date) == today.year
    ).scalar()

    if monthly_income is None:
        monthly_income = 0

    latest_budget = Budget.query.filter_by(
        user_id=current_user.user_id
    ).order_by(
        Budget.created_at.desc()
    ).first()

    if latest_budget:
        budget = latest_budget.budget_amount
    else:
        budget = 0

    savings = monthly_income - total_expense

    remaining_budget = budget - total_expense

    if budget > 0:
        budget_used = round((total_expense / budget) * 100, 2)
    else:
        budget_used = 0

    overspent = total_expense > budget

    recent_expenses = Expense.query.filter_by(
        user_id=current_user.user_id
    ).order_by(
        Expense.expense_date.desc()
    ).limit(5).all()

    for expense in recent_expenses:
        expense.converted_amount = convert_amount(expense.amount)

    goals = Goal.query.filter_by(
        user_id=current_user.user_id
    ).all()

    expense_count = Expense.query.filter_by(
        user_id=current_user.user_id
    ).count()

    income_count = Income.query.filter_by(
        user_id=current_user.user_id
    ).count()

    goal_count = Goal.query.filter_by(
        user_id=current_user.user_id
    ).count()

    monthly_expense = total_expense
    monthly_savings = savings
    budget_percentage = budget_used

    currency_symbol = get_currency_symbol()

    monthly_income = convert_amount(monthly_income)
    total_expense = convert_amount(total_expense)
    savings = convert_amount(savings)
    budget = convert_amount(budget)
    remaining_budget = convert_amount(remaining_budget)

    monthly_income = monthly_income
    monthly_expense = total_expense
    monthly_savings = savings

    # ---------------------------------
    # Expense Breakdown Filter
    # ---------------------------------

    expense_filter = request.args.get("expense_filter", "this_month")
    trend_filter = request.args.get("trend_filter", "this_month")

    from datetime import date, timedelta

    today = date.today()

    start_date = ""
    end_date = ""

    expense_query = Expense.query.filter(
        Expense.user_id == current_user.user_id
    )

    if expense_filter == "this_month":

        expense_query = expense_query.filter(
            db.extract("month", Expense.expense_date) == today.month,
            db.extract("year", Expense.expense_date) == today.year
        )

    elif expense_filter == "last_month":

        if today.month == 1:

            month = 12
            year = today.year - 1

        else:

            month = today.month - 1
            year = today.year

        expense_query = expense_query.filter(
            db.extract("month", Expense.expense_date) == month,
            db.extract("year", Expense.expense_date) == year
        )

    elif expense_filter == "last_3_months":

        three_months_ago = today - timedelta(days=90)

        expense_query = expense_query.filter(
            Expense.expense_date >= three_months_ago
        )

    elif expense_filter == "custom":

        start_date = request.args.get("start_date", "")
        end_date = request.args.get("end_date", "")

        if start_date and end_date:

            expense_query = expense_query.filter(
                Expense.expense_date.between(start_date, end_date)
            )

    category_data = db.session.query(
        Expense.category,
        db.func.sum(Expense.amount)
    ).filter(
        Expense.expense_id.in_(
            [e.expense_id for e in expense_query.all()]
        )
    ).group_by(
        Expense.category
    ).all()

    category_labels = [item[0] for item in category_data]
    category_amounts = [float(item[1]) for item in category_data]

    category_amounts = convert_list(category_amounts)

    # ---------------------------------
    # Monthly Trend
    # ---------------------------------

    trend_labels = []
    trend_income = []
    trend_expense = []

    if trend_filter == "this_month" or trend_filter == "last_month":

        trend_labels = ["Week 1", "Week 2", "Week 3", "Week 4"]

        trend_income = [0, 0, 0, 0]
        trend_expense = [0, 0, 0, 0]

        if trend_filter == "this_month":

            month = today.month
            year = today.year

        else:

            if today.month == 1:
                month = 12
                year = today.year - 1
            else:
                month = today.month - 1
                year = today.year

        incomes = Income.query.filter_by(
            user_id=current_user.user_id
        ).all()

        for income in incomes:

            if income.income_date.month == month and income.income_date.year == year:

                day = income.income_date.day

                if day <= 7:
                    trend_income[0] += float(income.amount)

                elif day <= 14:
                    trend_income[1] += float(income.amount)

                elif day <= 21:
                    trend_income[2] += float(income.amount)

                else:
                    trend_income[3] += float(income.amount)

        expenses = Expense.query.filter_by(
            user_id=current_user.user_id
        ).all()

        for expense in expenses:

            if expense.expense_date.month == month and expense.expense_date.year == year:

                day = expense.expense_date.day

                if day <= 7:
                    trend_expense[0] += float(expense.amount)

                elif day <= 14:
                    trend_expense[1] += float(expense.amount)

                elif day <= 21:
                    trend_expense[2] += float(expense.amount)

                else:
                    trend_expense[3] += float(expense.amount)

    elif trend_filter == "last_3_months":

        trend_labels = []
        trend_income = []
        trend_expense = []

        for i in range(2, -1, -1):

            month_date = today - timedelta(days=i * 30)

            month = month_date.month
            year = month_date.year

            trend_labels.append(month_date.strftime("%b"))

            income_total = db.session.query(
                db.func.sum(Income.amount)
            ).filter(
                Income.user_id == current_user.user_id,
                db.extract("month", Income.income_date) == month,
                db.extract("year", Income.income_date) == year
            ).scalar() or 0

            expense_total = db.session.query(
                db.func.sum(Expense.amount)
            ).filter(
                Expense.user_id == current_user.user_id,
                db.extract("month", Expense.expense_date) == month,
                db.extract("year", Expense.expense_date) == year
            ).scalar() or 0

            trend_income.append(float(income_total))
            trend_expense.append(float(expense_total))

    else:

        trend_labels = ["Income", "Expense"]

        income_total = 0
        expense_total = 0

        if start_date and end_date:

            income_total = db.session.query(
                db.func.sum(Income.amount)
            ).filter(
                Income.user_id == current_user.user_id,
                Income.income_date.between(start_date, end_date)
            ).scalar() or 0

            expense_total = db.session.query(
                db.func.sum(Expense.amount)
            ).filter(
                Expense.user_id == current_user.user_id,
                Expense.expense_date.between(start_date, end_date)
            ).scalar() or 0

        trend_income = [float(income_total), 0]
        trend_expense = [0, float(expense_total)]
    trend_income = convert_list(trend_income)
    trend_expense = convert_list(trend_expense)
    # -------------------------
    # NOTIFICATION COUNT
    # -------------------------

    

    notification_count = Notification.query.filter_by(
        user_id=current_user.user_id,
        is_read=False
    ).count()
   
    # -------------------------
    # INVESTMENT SUMMARY
    # -------------------------

    investments = Investment.query.filter_by(
        user_id=current_user.user_id
    ).all()

    # -------------------------
    # REFRESH CURRENT MARKET PRICES
    # -------------------------

    prices_updated = False

    for investment in investments:

        if investment.symbol:

            latest_price = get_stock_price(
                investment.symbol
            )

            if latest_price is not None:

                investment.current_price = latest_price

                prices_updated = True


    if prices_updated:

        db.session.commit()

    portfolio_value = 0
    total_investment = 0
    portfolio_profit = 0
    portfolio_return = 0

    for investment in investments:

        total_investment += investment.quantity * investment.buy_price

        portfolio_value += investment.quantity * investment.current_price

    portfolio_profit = portfolio_value - total_investment

    if total_investment > 0:

        portfolio_return = (
            portfolio_profit / total_investment
        ) * 100

    else:

        portfolio_return = 0

    return render_template(
        "dashboard.html",
        income=monthly_income,
        expense=total_expense,
        savings=savings,
        budget=budget,
        recent_expenses=recent_expenses,
        budget_used=budget_used,
        remaining_budget=remaining_budget,
        overspent=overspent,
        goals=goals,
        expense_count=expense_count,
        income_count=income_count,
        goal_count=goal_count,
        monthly_income=monthly_income,
        monthly_expense=monthly_expense,
        monthly_savings=monthly_savings,
        budget_percentage=budget_percentage,
        category_labels=category_labels,
        category_amounts=category_amounts,
        trend_labels=trend_labels,
        trend_income=trend_income,
        trend_expense=trend_expense,
        expense_filter=expense_filter,
        trend_filter=trend_filter,
        start_date=start_date,
        end_date=end_date,
        portfolio_value=portfolio_value,
        portfolio_profit=portfolio_profit,
        portfolio_return=portfolio_return,
        notification_count=notification_count,
        currency_symbol=currency_symbol
    )



# -------------------------
# EXPENSES
# -------------------------

@app.route("/expenses", methods=["GET", "POST"])
@login_required
def expenses():

    if request.method == "POST":

        new_expense = Expense(
            user_id=current_user.user_id,
            amount=convert_to_base_currency(float(request.form["amount"])),
            category=request.form["category"],
            payment_method=request.form["payment_method"],
            description=request.form["description"],
            expense_date=request.form["expense_date"]
        )

        db.session.add(new_expense)
        db.session.commit()

        generate_budget_notifications(current_user.user_id)
        generate_expense_notifications(current_user.user_id)
        generate_savings_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)
        generate_spending_insights(current_user.user_id)

        flash("Expense Added Successfully!", "success")

        return redirect("/expenses")

    expenses = Expense.query.filter_by(
        user_id=current_user.user_id
    ).order_by(
        Expense.expense_date.desc()
    ).all()

    for expense in expenses:
        expense.converted_amount = convert_amount(expense.amount)

    return render_template(
        "expenses.html",
        expenses=expenses,
        currency_symbol=get_currency_symbol()
    )

# -------------------------
# EDIT EXPENSE
# -------------------------

@app.route("/edit-expense/<int:expense_id>", methods=["GET", "POST"])
@login_required
def edit_expense(expense_id):

    expense = Expense.query.filter_by(
        expense_id=expense_id,
        user_id=current_user.user_id
    ).first_or_404()

    if request.method == "POST":

        expense.amount = convert_to_base_currency(float(request.form["amount"]))
        expense.category = request.form["category"]
        expense.payment_method = request.form["payment_method"]
        expense.description = request.form["description"]
        expense.expense_date = request.form["expense_date"]

        db.session.commit()

        generate_budget_notifications(current_user.user_id)
        generate_expense_notifications(current_user.user_id)
        generate_savings_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)
        generate_spending_insights(current_user.user_id)

        flash("Expense updated successfully!", "success")

        return redirect("/expenses")
    expense.converted_amount = convert_amount(expense.amount)
    return render_template(
        "edit_expense.html",
        expense=expense,
         currency_symbol=get_currency_symbol()
    )

# -------------------------
# DELETE EXPENSE
# -------------------------

@app.route("/delete-expense/<int:expense_id>")
@login_required
def delete_expense(expense_id):

    expense = Expense.query.filter_by(
        expense_id=expense_id,
        user_id=current_user.user_id
    ).first_or_404()

    db.session.delete(expense)

    db.session.commit()

    generate_budget_notifications(current_user.user_id)
    generate_savings_notifications(current_user.user_id)
    generate_financial_health_notifications(current_user.user_id)
    generate_spending_insights(current_user.user_id)

    flash("Expense deleted successfully!", "success")

    return redirect("/expenses")
# -------------------------
# BUDGETS
# -------------------------

@app.route("/budgets", methods=["GET", "POST"])
@login_required
def budgets():

    if request.method == "POST":

        budget_amount = request.form["budget_amount"]
        budget_month = request.form["budget_month"]
        budget_year = request.form["budget_year"]

        existing_budget = Budget.query.filter_by(
            user_id=current_user.user_id,
            budget_month=budget_month,
            budget_year=budget_year
        ).first()

        if existing_budget:

            existing_budget.budget_amount = budget_amount

        else:

            new_budget = Budget(
                user_id=current_user.user_id,
                budget_amount=budget_amount,
                budget_month=budget_month,
                budget_year=budget_year
            )

            db.session.add(new_budget)

        db.session.commit()

        generate_budget_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)

        flash("Budget saved successfully!", "success")

        return redirect("/budgets")

    budgets = Budget.query.filter_by(
        user_id=current_user.user_id
    ).order_by(
        Budget.budget_year.desc(),
        Budget.budget_month.desc()
    ).all()

    for budget in budgets:
        budget.converted_amount = convert_amount(budget.budget_amount)

    return render_template(
        "budgets.html",
        budgets=budgets,
        currency_symbol=get_currency_symbol()
    )


# -------------------------
# INCOME
# -------------------------

@app.route("/income", methods=["GET", "POST"])
@login_required
def income():

    if request.method == "POST":

        new_income = Income(
            user_id=current_user.user_id,
            source=request.form["source"],
            amount=request.form["amount"],
            income_date=request.form["income_date"],
            description=request.form["description"]
        )

        db.session.add(new_income)
        db.session.commit()

        generate_income_notifications(current_user.user_id)
        generate_savings_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)

        flash("Income added successfully!", "success")

        return redirect("/income")

    incomes = Income.query.filter_by(
        user_id=current_user.user_id
    ).order_by(
        Income.income_date.desc()
    ).all()

    for income in incomes:
        income.converted_amount = convert_amount(income.amount)

    return render_template(
        "income.html",
        incomes=incomes,
        currency_symbol=get_currency_symbol()
    )


# -------------------------
# GOALS
# -------------------------

@app.route("/goals", methods=["GET", "POST"])
@login_required
def goals():

    if request.method == "POST":

        new_goal = Goal(
            user_id=current_user.user_id,
            goal_name=request.form["goal_name"],
            target_amount=request.form["target_amount"],
            saved_amount=request.form["saved_amount"] or 0,
            target_date=request.form["target_date"]
        )

        db.session.add(new_goal)
        db.session.commit()

        generate_goal_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)

        flash("Goal created successfully!", "success")

        return redirect("/goals")

    goals = Goal.query.filter_by(
        user_id=current_user.user_id
    ).all()

    for goal in goals:

        goal.converted_target = convert_amount(goal.target_amount)
        goal.converted_saved = convert_amount(goal.saved_amount)

        if goal.target_amount > 0:

            goal.progress = round(
                (goal.saved_amount / goal.target_amount) * 100,
                1
            )

        else:

            goal.progress = 0

        if goal.progress >= 100:

            goal.status = "Completed"

        elif goal.progress >= 50:

            goal.status = "On Track"

        else:

            goal.status = "Needs Attention"

    
    return render_template(
        "goals.html",
        goals=goals,
        currency_symbol=get_currency_symbol()
    )

# -------------------------
# UPDATE GOAL
# -------------------------

@app.route(
    "/update-goal/<int:goal_id>",
    methods=["GET", "POST"]
)
@login_required
def update_goal(goal_id):

    goal = Goal.query.filter_by(
        goal_id=goal_id,
        user_id=current_user.user_id
    ).first_or_404()

    if request.method == "POST":

        goal.goal_name = request.form["goal_name"]
        goal.target_amount = float(
            request.form["target_amount"]
        )
        goal.saved_amount = float(
            request.form["saved_amount"]
        )
        goal.target_date = request.form["target_date"]

        db.session.commit()

        generate_goal_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)

        flash(
            "Goal updated successfully!",
            "success"
        )

        return redirect(
            url_for("goals")
        )

    goal.converted_target = convert_amount(goal.target_amount)
    goal.converted_saved = convert_amount(goal.saved_amount)
    return render_template(
        "update_goal.html",
        goal=goal,
        currency_symbol=get_currency_symbol()
    )

# -------------------------
# SAVE MORE TO GOAL
# -------------------------

@app.route(
    "/save-goal/<int:goal_id>",
    methods=["GET", "POST"]
)
@login_required
def save_goal(goal_id):

    goal = Goal.query.filter_by(
        goal_id=goal_id,
        user_id=current_user.user_id
    ).first_or_404()

    if request.method == "POST":

        amount = float(
            request.form["amount"]
        )

        if amount <= 0:

            flash(
                "Amount must be greater than zero.",
                "danger"
            )

            return redirect(
                url_for(
                    "save_goal",
                    goal_id=goal.goal_id
                )
            )

        goal.saved_amount += amount

        db.session.commit()

        generate_goal_notifications(current_user.user_id)
        generate_savings_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)

        flash(
            "Savings added successfully!",
            "success"
        )

        return redirect(
            url_for("goals")
        )

    goal.converted_saved = convert_amount(goal.saved_amount)

    return render_template(
        "save_goal.html",
        goal=goal,
        currency_symbol=get_currency_symbol()
    )

# -------------------------
# INVESTMENTS
# -------------------------

@app.route("/investments")
@login_required
def investments():

    # -------------------------
    # ACTIVE HOLDINGS
    # -------------------------

    investments = Investment.query.filter(
        Investment.user_id == current_user.user_id,
        Investment.quantity > 0
    ).order_by(
        Investment.purchase_date.desc()
    ).all()

    # -------------------------
    # REFRESH MARKET PRICES
    # -------------------------

    refresh_investment_prices(
        investments,
        db
    )

    # -------------------------
    # PORTFOLIO CALCULATIONS
    # -------------------------

    total_investment = 0
    current_value = 0
    unrealized_profit = 0

    for investment in investments:

        invested = (
            investment.quantity
            * investment.buy_price
        )

        current = (
            investment.quantity
            * investment.current_price
        )

        total_investment += invested

        current_value += current

        unrealized_profit += (
            current - invested
        )

    # -------------------------
    # REALIZED PROFIT
    # TRANSACTION COST BASIS
    # -------------------------

    all_transactions = db.session.query(
        InvestmentTransaction
    ).join(
        Investment,
        InvestmentTransaction.investment_id
        == Investment.investment_id
    ).filter(
        InvestmentTransaction.user_id
        == current_user.user_id
    ).order_by(
        InvestmentTransaction.transaction_date.asc(),
        InvestmentTransaction.transaction_id.asc()
    ).all()

    transaction_groups = {}

    for transaction in all_transactions:

        if transaction.investment_id not in transaction_groups:

            transaction_groups[
                transaction.investment_id
            ] = []

        transaction_groups[
            transaction.investment_id
        ].append(transaction)

    realized_profit = 0

    for investment_id, transactions in transaction_groups.items():

        held_quantity = 0
        average_cost = 0

        for transaction in transactions:

            if transaction.transaction_type == "BUY":

                old_cost = (
                    held_quantity * average_cost
                )

                new_cost = (
                    transaction.quantity
                    * transaction.price
                )

                held_quantity += transaction.quantity

                if held_quantity > 0:

                    average_cost = (
                        old_cost + new_cost
                    ) / held_quantity

            elif transaction.transaction_type == "SELL":

                realized_profit += (
                    transaction.price
                    - average_cost
                ) * transaction.quantity

                held_quantity -= transaction.quantity

                if held_quantity <= 0:

                    held_quantity = 0
                    average_cost = 0

    # -------------------------
    # TOTAL PROFIT & RETURN
    # -------------------------

    total_profit = (
        realized_profit
        + unrealized_profit
    )

    total_buy_amount = 0

    for transaction in all_transactions:

        if transaction.transaction_type == "BUY":

            total_buy_amount += (
                transaction.quantity
                * transaction.price
            )


    if total_buy_amount > 0:

        return_percentage = (
            total_profit
            / total_buy_amount
        ) * 100

    else:

        return_percentage = 0

    # -------------------------
    # ASSET ALLOCATION
    # -------------------------

    allocation = {}

    for investment in investments:

        value = (
            investment.quantity
            * investment.current_price
        )

        allocation[
            investment.investment_type
        ] = allocation.get(
            investment.investment_type,
            0
        ) + value

    allocation_labels = list(
        allocation.keys()
    )

    allocation_values = list(
        allocation.values()
    )
    # -------------------------
    # PORTFOLIO PERFORMANCE
    # HISTORICAL VALUE
    # -------------------------

    performance_labels = []
    performance_values = []

    performance_transactions = db.session.query(
        InvestmentTransaction,
        Investment
    ).join(
        Investment,
        InvestmentTransaction.investment_id
        == Investment.investment_id
    ).filter(
        InvestmentTransaction.user_id
        == current_user.user_id
    ).order_by(
        InvestmentTransaction.transaction_date.asc(),
        InvestmentTransaction.transaction_id.asc()
    ).all()


    transactions_by_date = {}

    for transaction, investment in performance_transactions:

        transaction_date = transaction.transaction_date

        if transaction_date not in transactions_by_date:
            transactions_by_date[transaction_date] = []

        transactions_by_date[transaction_date].append(
            (transaction, investment)
        )


    holding_quantities = {}


    for transaction_date in sorted(transactions_by_date.keys()):

        for transaction, investment in transactions_by_date[
            transaction_date
        ]:

            investment_id = investment.investment_id

            if investment_id not in holding_quantities:

                holding_quantities[investment_id] = {
                    "quantity": 0,
                    "symbol": investment.symbol
                }

            if transaction.transaction_type == "BUY":

                holding_quantities[
                    investment_id
                ]["quantity"] += transaction.quantity

            elif transaction.transaction_type == "SELL":

                holding_quantities[
                    investment_id
                ]["quantity"] -= transaction.quantity


        portfolio_value = 0


        for holding in holding_quantities.values():

            if holding["quantity"] <= 0:
                continue

            historical_price = get_historical_price(
                holding["symbol"],
                transaction_date
            )

            if historical_price is not None:

                portfolio_value += (
                    holding["quantity"]
                    * historical_price
                )


        performance_labels.append(
            transaction_date.strftime("%d-%b-%Y")
        )

        performance_values.append(
            round(portfolio_value, 2)
        )

    # -------------------------
    # TOP HOLDINGS
    # -------------------------

    top_holdings = sorted(
        investments,
        key=lambda x:
            x.quantity * x.current_price,
        reverse=True
    )[:5]
    market_data = get_market_overview()

    # -------------------------
    # PORTFOLIO SUMMARY
    # -------------------------

    total_holdings = len(investments)

    asset_types = len(
        set(
            investment.investment_type
            for investment in investments
        )
    )

    largest_holding = None
    best_performer = None
    worst_performer = None

    if investments:

        largest_holding = max(
            investments,
            key=lambda x: x.quantity * x.current_price
        )

        best_performer = max(
            investments,
            key=lambda x: (
                (
                    (
                        x.current_price - x.buy_price
                    )
                    / x.buy_price
                ) if x.buy_price > 0 else 0
            )
        )

        worst_performer = min(
            investments,
            key=lambda x: (
                (
                    (
                        x.current_price - x.buy_price
                    )
                    / x.buy_price
                ) if x.buy_price > 0 else 0
            )
        )

    total_investment = convert_amount(total_investment)
    current_value = convert_amount(current_value)
    realized_profit = convert_amount(realized_profit)
    unrealized_profit = convert_amount(unrealized_profit)
    total_profit = convert_amount(total_profit)

    allocation_values = convert_list(allocation_values)
    performance_values = convert_list(performance_values)

    for investment in top_holdings:
        investment.converted_current_value = convert_amount(
            investment.quantity * investment.current_price
        )

    return render_template(
        "investments.html",
        investments=investments,
        total_investment=total_investment,
        current_value=current_value,
        total_profit=total_profit,
        return_percentage=return_percentage,
        allocation_labels=allocation_labels,
        allocation_values=allocation_values,
        performance_labels=performance_labels,
        performance_values=performance_values,
        top_holdings=top_holdings,
        realized_profit=realized_profit,
        unrealized_profit=unrealized_profit,
        market_data=market_data,
        total_holdings=total_holdings,
        asset_types=asset_types,
        largest_holding=largest_holding,
        best_performer=best_performer,
        worst_performer=worst_performer,
        currency_symbol=get_currency_symbol()
    )

# -------------------------
# ADD INVESTMENT
# -------------------------

@app.route("/add-investment", methods=["GET", "POST"])
@login_required
def add_investment():

    if request.method == "POST":

        investment_name = request.form["investment_name"].strip()
        investment_type = request.form["investment_type"]

        symbol = request.form.get("symbol", "").strip().upper()

        quantity = float(request.form["quantity"])
        buy_price = float(request.form["buy_price"])
        purchase_date = request.form["purchase_date"]

        # VALIDATE QUANTITY AND BUY PRICE

        if quantity <= 0 or buy_price <= 0:

            flash(
                "Quantity and buy price must be greater than zero.",
                "danger"
            )

            return redirect(url_for("add_investment"))

        # VALIDATE SYMBOL

        if not symbol:

            flash(
                "Stock symbol is required.",
                "danger"
            )

            return redirect(url_for("add_investment"))

        current_price = get_stock_price(symbol)

        if current_price is None:

            flash(
                "Invalid stock symbol or market price unavailable.",
                "danger"
            )

            return redirect(url_for("add_investment"))

        # CREATE INVESTMENT

        investment = Investment(

            user_id=current_user.user_id,

            investment_name=investment_name,

            investment_type=investment_type,

            symbol=symbol,

            quantity=quantity,

            buy_price=buy_price,

            current_price=current_price,

            purchase_date=purchase_date

        )

        db.session.add(investment)

        db.session.flush()

        # CREATE BUY TRANSACTION

        buy_transaction = InvestmentTransaction(

            user_id=current_user.user_id,

            investment_id=investment.investment_id,

            transaction_type="BUY",

            quantity=investment.quantity,

            price=investment.buy_price,

            transaction_date=investment.purchase_date

        )

        db.session.add(buy_transaction)

        db.session.commit()

        generate_investment_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)

        flash(
            "Investment added with latest market price!",
            "success"
        )

        return redirect(url_for("holdings"))

    return render_template("add_investment.html")
# -------------------------
# EDIT INVESTMENT
# -------------------------

@app.route(
    "/edit-investment/<int:investment_id>",
    methods=["GET", "POST"]
)
@login_required
def edit_investment(investment_id):

    investment = Investment.query.filter_by(
        investment_id=investment_id,
        user_id=current_user.user_id
    ).first_or_404()

    if request.method == "POST":

        investment.investment_name = request.form[
            "investment_name"
        ]

        investment.investment_type = request.form[
            "investment_type"
        ]

        db.session.commit()

        generate_investment_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)

        flash(
            "Investment details updated successfully!",
            "success"
        )

        return redirect(
            url_for("holdings")
        )
    investment.converted_buy_price = convert_amount(investment.buy_price)
    return render_template(
        "edit_investment.html",
        investment=investment,
        currency_symbol=get_currency_symbol()
    )

# -------------------------
# DELETE INVESTMENT
# -------------------------
@app.route("/delete-investment/<int:investment_id>")
@login_required
def delete_investment(investment_id):

    investment = Investment.query.filter_by(
        investment_id=investment_id,
        user_id=current_user.user_id
    ).first_or_404()

    InvestmentTransaction.query.filter_by(
        investment_id=investment.investment_id,
        user_id=current_user.user_id
    ).delete(synchronize_session=False)

    db.session.delete(investment)

    db.session.commit()

    generate_investment_notifications(current_user.user_id)
    generate_financial_health_notifications(current_user.user_id)

    flash(
        "Investment and related transactions deleted successfully!",
        "success"
    )

    return redirect(url_for("holdings"))
# -------------------------
# HOLDINGS
# -------------------------

@app.route("/holdings")
@login_required
def holdings():

    investments = Investment.query.filter(
        Investment.user_id == current_user.user_id,
        Investment.quantity > 0
    ).order_by(
        Investment.purchase_date.desc()
    ).all()

    refresh_investment_prices(
        investments,
        db
    )

    risk_map = {

        "Stock": "High",

        "Crypto": "Very High",

        "Mutual Fund": "Medium",

        "ETF": "Medium",

        "Gold": "Low",

        "Bond": "Low",

        "Fixed Deposit": "Very Low"

    }

    for investment in investments:

        investment.risk_level = risk_map.get(
            investment.investment_type,
            "Unknown"
        )

    for investment in investments:
        investment.converted_buy_price = convert_amount(investment.buy_price)
        investment.converted_current_price = convert_amount(investment.current_price)

        investment.converted_invested = convert_amount(
            investment.quantity * investment.buy_price
        )

        investment.converted_current_value = convert_amount(
            investment.quantity * investment.current_price
        )

        investment.converted_profit = convert_amount(
            (investment.quantity * investment.current_price)
            - (investment.quantity * investment.buy_price)
        )

    return render_template(
        "holdings.html",
        investments=investments,
        currency_symbol=get_currency_symbol()
    )

# -------------------------
# TRANSACTIONS
# -------------------------

@app.route("/transactions")
@login_required
def transactions():

    transaction_records = db.session.query(
        InvestmentTransaction,
        Investment
    ).join(
        Investment,
        InvestmentTransaction.investment_id
        == Investment.investment_id
    ).filter(
        InvestmentTransaction.user_id
        == current_user.user_id
    ).order_by(
        InvestmentTransaction.transaction_date.asc(),
        InvestmentTransaction.transaction_id.asc()
    ).all()

    holdings_cost = {}

    transaction_data = []

    for transaction, investment in transaction_records:

        investment_id = transaction.investment_id

        if investment_id not in holdings_cost:

            holdings_cost[investment_id] = {
                "quantity": 0,
                "average_cost": 0
            }

        holding = holdings_cost[investment_id]

        realized_profit = None

        if transaction.transaction_type == "BUY":

            old_cost = (
                holding["quantity"]
                * holding["average_cost"]
            )

            new_cost = (
                transaction.quantity
                * transaction.price
            )

            holding["quantity"] += transaction.quantity

            if holding["quantity"] > 0:

                holding["average_cost"] = (
                    old_cost + new_cost
                ) / holding["quantity"]

        elif transaction.transaction_type == "SELL":

            realized_profit = (
                transaction.price
                - holding["average_cost"]
            ) * transaction.quantity

            holding["quantity"] -= transaction.quantity

            if holding["quantity"] <= 0:

                holding["quantity"] = 0
                holding["average_cost"] = 0

        transaction_data.append({
            "transaction": transaction,
            "investment": investment,
            "realized_profit": realized_profit
        })

    transaction_data.reverse()

    for item in transaction_data:

        transaction = item["transaction"]

        transaction.converted_price = convert_amount(
            transaction.price
        )

        item["converted_total_amount"] = convert_amount(
            transaction.quantity * transaction.price
        )

        if item["realized_profit"] is not None:

            item["converted_realized_profit"] = convert_amount(
                item["realized_profit"]
            )

    return render_template(
        "transactions.html",
        transactions=transaction_data,
        currency_symbol=get_currency_symbol()
    )

# -------------------------
# SELL INVESTMENT
# -------------------------
@app.route(
    "/sell-investment/<int:investment_id>",
    methods=["GET", "POST"]
)
@login_required
def sell_investment(investment_id):

    from datetime import date

    investment = Investment.query.filter_by(
        investment_id=investment_id,
        user_id=current_user.user_id
    ).first_or_404()

    if request.method == "POST":

        try:
            sell_quantity = float(
                request.form["quantity"]
            )

        except (ValueError, TypeError):

            flash(
                "Enter a valid sell quantity.",
                "danger"
            )

            return redirect(
                url_for(
                    "sell_investment",
                    investment_id=investment_id
                )
            )

        # QUANTITY VALIDATION

        if sell_quantity <= 0:

            flash(
                "Sell quantity must be greater than zero.",
                "danger"
            )

            return redirect(
                url_for(
                    "sell_investment",
                    investment_id=investment_id
                )
            )

        if sell_quantity > investment.quantity:

            flash(
                "You cannot sell more than your current holdings.",
                "danger"
            )

            return redirect(
                url_for(
                    "sell_investment",
                    investment_id=investment_id
                )
            )

        # SYMBOL VALIDATION

        if not investment.symbol:

            flash(
                "Investment symbol is unavailable.",
                "danger"
            )

            return redirect(
                url_for("holdings")
            )

        # FETCH LIVE SELL PRICE

        latest_price = get_stock_price(
            investment.symbol
        )

        if latest_price is None:

            flash(
                "Unable to fetch latest market price. Please try again.",
                "danger"
            )

            return redirect(
                url_for(
                    "sell_investment",
                    investment_id=investment_id
                )
            )

        # CREATE SELL TRANSACTION

        sell_transaction = InvestmentTransaction(

            user_id=current_user.user_id,

            investment_id=investment.investment_id,

            transaction_type="SELL",

            quantity=sell_quantity,

            price=latest_price,

            transaction_date=date.today()

        )

        # UPDATE HOLDING

        investment.quantity = round(
            investment.quantity - sell_quantity,
            6
        )

        investment.current_price = latest_price

        db.session.add(sell_transaction)

        db.session.commit()

        generate_investment_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)

        flash(
            f"Sold {sell_quantity:g} units at ₹{latest_price:.2f} successfully!",
            "success"
        )

        return redirect(
            url_for("holdings")
        )

    investment.converted_current_price = convert_amount(investment.current_price)

    return render_template(
        "sell_investment.html",
        investment=investment,
        currency_symbol=get_currency_symbol()
    )
# -------------------------
# BUY MORE INVESTMENT
# -------------------------
@app.route(
    "/buy-more/<int:investment_id>",
    methods=["GET", "POST"]
)
@login_required
def buy_more(investment_id):

    from datetime import date

    investment = Investment.query.filter_by(
        investment_id=investment_id,
        user_id=current_user.user_id
    ).first_or_404()

    if request.method == "POST":

        try:

            buy_quantity = float(
                request.form["quantity"]
            )

            buy_price = float(
                request.form["buy_price"]
            )

        except (ValueError, TypeError):

            flash(
                "Enter valid quantity and buy price.",
                "danger"
            )

            return redirect(
                url_for(
                    "buy_more",
                    investment_id=investment_id
                )
            )

        # VALIDATION

        if buy_quantity <= 0 or buy_price <= 0:

            flash(
                "Quantity and buy price must be greater than zero.",
                "danger"
            )

            return redirect(
                url_for(
                    "buy_more",
                    investment_id=investment_id
                )
            )

        # CALCULATE WEIGHTED AVERAGE COST

        old_quantity = investment.quantity

        old_invested_amount = (
            old_quantity
            * investment.buy_price
        )

        new_invested_amount = (
            buy_quantity
            * buy_price
        )

        total_quantity = (
            old_quantity
            + buy_quantity
        )

        average_buy_price = (
            old_invested_amount
            + new_invested_amount
        ) / total_quantity

        # CREATE BUY TRANSACTION

        buy_transaction = InvestmentTransaction(

            user_id=current_user.user_id,

            investment_id=investment.investment_id,

            transaction_type="BUY",

            quantity=buy_quantity,

            price=buy_price,

            transaction_date=date.today()

        )

        # UPDATE HOLDING

        investment.quantity = round(
            total_quantity,
            6
        )

        investment.buy_price = round(
            average_buy_price,
            2
        )

        # REFRESH CURRENT MARKET PRICE

        if investment.symbol:

            latest_price = get_stock_price(
                investment.symbol
            )

            if latest_price is not None:

                investment.current_price = latest_price

        db.session.add(buy_transaction)

        db.session.commit()

        generate_investment_notifications(current_user.user_id)
        generate_financial_health_notifications(current_user.user_id)

        flash(
            f"Purchased {buy_quantity:g} additional units at ₹{buy_price:.2f}!",
            "success"
        )

        return redirect(
            url_for("holdings")
        )
    investment.converted_buy_price = convert_amount(investment.buy_price)

    return render_template(
        "buy_more.html",
        investment=investment,
        currency_symbol=get_currency_symbol()
    )
# -------------------------
# ASSET ALLOCATION
# -------------------------

@app.route("/asset-allocation")
@login_required
def asset_allocation():

    investments = Investment.query.filter(
        Investment.user_id == current_user.user_id,
        Investment.quantity > 0
    ).all()

    refresh_investment_prices(
        investments,
        db
    )

    allocation = {}

    total_portfolio_value = 0

    for investment in investments:

        current_price = investment.current_price or 0

        value = (
            investment.quantity
            * current_price
        )

        total_portfolio_value += value

        allocation[
            investment.investment_type
        ] = allocation.get(
            investment.investment_type,
            0
        ) + value

    labels = list(allocation.keys())

    values = list(allocation.values())

    allocation_percentages = {}

    for investment_type, value in allocation.items():

        if total_portfolio_value > 0:

            percentage = (
                value / total_portfolio_value
            ) * 100

        else:

            percentage = 0

        allocation_percentages[
            investment_type
        ] = percentage

    # -------------------------
    # PORTFOLIO INSIGHT
    # -------------------------

    portfolio_insight = ""

    if allocation_percentages:

        highest_asset = max(
            allocation_percentages,
            key=allocation_percentages.get
        )

        highest_percentage = allocation_percentages[
            highest_asset
        ]

        if highest_percentage >= 80:

            portfolio_insight = (
                f"⚠ Your portfolio is highly concentrated in "
                f"{highest_asset} ({highest_percentage:.1f}%). "
                f"Consider diversifying into other asset classes."
            )

        elif highest_percentage >= 60:

            portfolio_insight = (
                f"Your portfolio is mostly invested in "
                f"{highest_asset} ({highest_percentage:.1f}%). "
                f"A little more diversification may reduce risk."
            )

        else:

            portfolio_insight = (
                "✅ Your portfolio appears to be well diversified."
            )

    else:

        portfolio_insight = (
            "Add investments to receive diversification insights."
        )

    total_portfolio_value = convert_amount(total_portfolio_value)

    values = convert_list(values)

    converted_allocation = {}

    for key, value in allocation.items():
        converted_allocation[key] = convert_amount(value)

    return render_template(
        "asset_allocation.html",
        labels=labels,
        allocation=converted_allocation,
        values=values,
        total_portfolio_value=total_portfolio_value,
        currency_symbol=get_currency_symbol(),
        allocation_percentages=allocation_percentages,
        portfolio_insight=portfolio_insight,
    )
# -------------------------
# PERFORMANCE
# -------------------------

@app.route("/performance")
@login_required
def performance():

    investments = Investment.query.filter(
        Investment.user_id == current_user.user_id,
        Investment.quantity > 0
    ).all()

    refresh_investment_prices(
        investments,
        db
    )

    labels = []

    portfolio_values = []

    invested_values = []

    for investment in investments:

        labels.append(
            investment.investment_name
        )

        invested_values.append(
            investment.quantity
            * investment.buy_price
        )

        portfolio_values.append(
            investment.quantity
            * (investment.current_price or 0)
        )

    total_investment = sum(
        invested_values
    )

    current_value = sum(
        portfolio_values
    )

    unrealized_profit = (
        current_value - total_investment
    )

    if total_investment > 0:

        return_percent = (
            unrealized_profit
            / total_investment
        ) * 100

    else:

        return_percent = 0

    total_investment = convert_amount(total_investment)
    current_value = convert_amount(current_value)
    unrealized_profit = convert_amount(unrealized_profit)

    invested_values = convert_list(invested_values)
    portfolio_values = convert_list(portfolio_values)

    return render_template(
        "performance.html",
        labels=labels,
        portfolio_values=portfolio_values,
        invested_values=invested_values,
        total_investment=total_investment,
        current_value=current_value,
        profit=unrealized_profit,
        return_percent=return_percent,
        currency_symbol=get_currency_symbol()
    )
# -------------------------
# ANALYTICS
# -------------------------

@app.route("/analytics")
@login_required
def analytics():

    selected_year = request.args.get("year", datetime.now().year, type=int)

    incomes = Income.query.filter_by(
        user_id=current_user.user_id
    ).all()

    expenses = Expense.query.filter_by(
        user_id=current_user.user_id
    ).all()

    budgets = Budget.query.filter_by(
        user_id=current_user.user_id
    ).all()

    investments = Investment.query.filter_by(
        user_id=current_user.user_id
    ).all()

    goals = Goal.query.filter_by(
        user_id=current_user.user_id
    ).all()

    total_income = sum(
        income.amount
        for income in incomes
    )

    total_expense = sum(
        expense.amount
        for expense in expenses
    )

    # -------------------------
    # SPENDING PATTERN
    # -------------------------

    expense_categories = {}

    for expense in expenses:

        # Selected year expenses matrame include cheyyi
        if expense.expense_date.year == selected_year:

            category = expense.category

            expense_categories[category] = (
                expense_categories.get(category, 0)
                + expense.amount
            )

    expense_labels = list(expense_categories.keys())
    expense_values = list(expense_categories.values())

    highest_category = None
    highest_amount = 0

    if expense_categories:
        highest_category = max(expense_categories, key=expense_categories.get)
        highest_amount = expense_categories[highest_category]
    
    # -------------------------
    # CASH FLOW TREND
    # ------------------------- 

    months = [
        "Jan", "Feb", "Mar", "Apr",
        "May", "Jun", "Jul", "Aug",
        "Sep", "Oct", "Nov", "Dec"
    ]

    income_data = [0] * 12
    expense_data = [0] * 12
    saving_data = [0] * 12

    for income in incomes:
        if income.income_date.year == selected_year:
            month = income.income_date.month - 1
            income_data[month] += income.amount
    for expense in expenses:
        if expense.expense_date.year == selected_year:
            month = expense.expense_date.month - 1
            expense_data[month] += expense.amount

    for i in range(12):

        saving_data[i] = (
            income_data[i]
            - expense_data[i]
        )

    total_budget = sum(
        budget.budget_amount
        for budget in budgets
    )

    total_investment = sum(
        investment.quantity
        * investment.current_price
        for investment in investments
    )

    completed_goals = sum(
        1
        for goal in goals
        if goal.saved_amount >= goal.target_amount
    )

    

    health_score = 100

    # Savings Rate
    if total_income > 0:

        savings_rate = ((total_income - total_expense) / total_income) * 100

        if savings_rate < 10:
            health_score -= 25
        elif savings_rate < 20:
            health_score -= 15
        elif savings_rate < 30:
            health_score -= 5


    # Expense Ratio
    expense_ratio = (total_expense / total_income) * 100 if total_income else 100

    if expense_ratio > 80:
        health_score -= 20
    elif expense_ratio > 60:
        health_score -= 10


    # Budget
    if total_budget > 0 and total_expense > total_budget:
        health_score -= 15


    # Investments
    if total_investment == 0:
        health_score -= 10


    # Goals
    if completed_goals == 0 and len(goals) > 0:
        health_score -= 10


    health_score = max(0, min(100, health_score))

    if health_score >= 90:
        health_status = "Excellent"
    elif health_score >= 75:
        health_status = "Good"
    elif health_score >= 60:
        health_status = "Average"
    elif health_score >= 40:
        health_status = "Needs Improvement"
    else:
        health_status = "Poor"

    # ============================
    # Financial Insights
    # ============================

    financial_insights = []
    budget_recommendations = []
    alerts = []

    savings = total_income - total_expense
    savings_rate = (savings / total_income * 100) if total_income > 0 else 0
    expense_ratio = (total_expense / total_income * 100) if total_income > 0 else 0
    investment_ratio = (total_investment / total_income * 100) if total_income > 0 else 0

    if savings_rate >= 30:
        financial_insights.append(
            "Excellent! You are saving more than 30% of your income."
        )
    elif savings_rate >= 20:
        financial_insights.append(
            "Your savings rate is healthy. Keep maintaining it."
        )
    else:
        financial_insights.append(
            "Try increasing your savings to at least 20% of your income."
        )

    if investment_ratio >= 20:
        financial_insights.append(
            "Your investment portfolio is growing well."
        )
    elif total_investment > 0:
        financial_insights.append(
            "Consider increasing your investments gradually."
        )

    if completed_goals == len(goals) and len(goals) > 0:
        financial_insights.append(
            "Congratulations! All your financial goals are completed."
        )

    # ============================
    # Budget Recommendations
    # ============================

    if expense_ratio > 70:
        budget_recommendations.append(
            "Your expenses are more than 70% of your income. Try reducing non-essential spending."
        )

    if savings_rate < 20:
        budget_recommendations.append(
            "Aim to save at least 20% of your monthly income."
        )

    if total_budget > 0 and total_expense > total_budget:
        budget_recommendations.append(
            "You have exceeded your planned budget. Review your highest spending category."
        )

    if total_investment == 0:
        budget_recommendations.append(
            "Consider starting investments to build long-term wealth."
        )

    if not budget_recommendations:
        budget_recommendations.append(
            "Excellent! Your budgeting habits are on track. Keep maintaining them."
        )

    monthly_savings = convert_amount(total_income - total_expense)
    projected_balance = convert_amount(
        (total_income - total_expense) + total_investment
    )

    total_income = convert_amount(total_income)
    total_expense = convert_amount(total_expense)
    total_investment = convert_amount(total_investment)

    expense_values = convert_list(expense_values)

    income_data = convert_list(income_data)
    expense_data = convert_list(expense_data)
    saving_data = convert_list(saving_data)

    return render_template(

        "analytics.html",

        budget_recommendations=budget_recommendations,
        financial_insights=financial_insights,

        highest_category=highest_category,
        highest_amount=highest_amount,

        selected_year=selected_year,

        total_income=total_income,

        total_expense=total_expense,

        total_budget=total_budget,

        total_investment=total_investment,

        total_goals=len(goals),

        completed_goals=completed_goals,
        health_score=health_score,

        health_status=health_status,

        expense_labels=expense_labels,

        expense_values=expense_values,

        months=months,

        income_data=income_data,

        expense_data=expense_data,

        saving_data=saving_data,

        projected_balance=projected_balance,

        currency_symbol=get_currency_symbol()

    )


# -------------------------
# REPORTS
# -------------------------

@app.route("/reports")
@login_required
def reports():
    return render_template("reports.html")



def get_report_data(user_id, period, from_date=None, to_date=None):

    today = datetime.today()

    start_date = None
    end_date = None

    if period == "this_month":

        start_date = today.replace(day=1)
        end_date = today

    elif period == "last_month":

        first_this_month = today.replace(day=1)
        end_date = first_this_month - timedelta(days=1)
        start_date = end_date.replace(day=1)

    elif period == "last_3_months":

        start_date = today - timedelta(days=90)
        end_date = today

    elif period == "this_year":

        start_date = today.replace(month=1, day=1)
        end_date = today

    elif period == "custom" and from_date and to_date:

        start_date = datetime.strptime(from_date, "%Y-%m-%d")
        end_date = datetime.strptime(to_date, "%Y-%m-%d")

    income_query = Income.query.filter_by(user_id=user_id)

    if start_date and end_date:
        income_query = income_query.filter(
            Income.income_date.between(start_date.date(), end_date.date())
        )

    incomes = income_query.all()

    expense_query = Expense.query.filter_by(user_id=user_id)

    if start_date and end_date:
        expense_query = expense_query.filter(
            Expense.expense_date.between(start_date.date(), end_date.date())
        )

    expenses = expense_query.all()

    budget_query = Budget.query.filter_by(user_id=user_id)

    if start_date and end_date:
        budget_query = budget_query.filter(
            Budget.created_at.between(start_date, end_date)
        )

    budgets = budget_query.all()

    goal_query = Goal.query.filter_by(user_id=user_id)

    if start_date and end_date:
        goal_query = goal_query.filter(
            Goal.created_at.between(start_date, end_date)
        )

    goals = goal_query.all()

    investment_query = Investment.query.filter_by(user_id=user_id)

    if start_date and end_date:
        investment_query = investment_query.filter(
            Investment.purchase_date.between(start_date.date(), end_date.date())
        )

    investments = investment_query.all()

    notification_query = Notification.query.filter_by(user_id=user_id)

    if start_date and end_date:
        notification_query = notification_query.filter(
            Notification.created_at.between(start_date, end_date)
        )

    notifications = notification_query.order_by(
        Notification.created_at.desc()
    ).all()

    total_income = sum(i.amount for i in incomes)

    total_expense = sum(e.amount for e in expenses)

    total_savings = total_income - total_expense

    total_investment = sum(
        inv.quantity * inv.current_price
        for inv in investments
    )

    net_worth = total_savings + total_investment

    savings_rate = (
        round((total_savings / total_income) * 100, 2)
        if total_income > 0 else 0
    )

    health_score = 80

    return {
        "incomes": incomes,
        "expenses": expenses,
        "budgets": budgets,
        "goals": goals,
        "investments": investments,
        "notifications": notifications,
        "total_income": total_income,
        "total_expense": total_expense,
        "total_savings": total_savings,
        "total_investment": total_investment,
        "net_worth": net_worth,
        "savings_rate": savings_rate,
        "health_score": health_score,
    }

@app.route("/reports/preview")
@login_required
def report_preview():

    period = request.args.get("period", "this_month")
    from_date = request.args.get("from")
    to_date = request.args.get("to")

    report = get_report_data(
        current_user.user_id,
        period,
        from_date,
        to_date
    )

    # Convert summary values
    report["total_income"] = convert_amount(report["total_income"])
    report["total_expense"] = convert_amount(report["total_expense"])
    report["total_savings"] = convert_amount(report["total_savings"])

    # Income
    for income in report["incomes"]:
        income.converted_amount = convert_amount(income.amount)

    # Expense
    for expense in report["expenses"]:
        expense.converted_amount = convert_amount(expense.amount)

    # Budget
    for budget in report["budgets"]:
        budget.converted_amount = convert_amount(budget.budget_amount)

    # Goal
    for goal in report["goals"]:
        goal.converted_target = convert_amount(goal.target_amount)
        goal.converted_saved = convert_amount(goal.saved_amount)

    # Investment
    for investment in report["investments"]:
        investment.converted_buy_price = convert_amount(investment.buy_price)
        investment.converted_current_price = convert_amount(investment.current_price)
        investment.converted_current_value = convert_amount(
            investment.quantity * investment.current_price
        )

    report["currency_symbol"] = get_currency_symbol()
    return render_template(
        "report_preview.html",
        **report
    )

@app.route("/reports/export/pdf")
@login_required
def export_pdf():

    report = get_report_data(current_user.user_id)

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    elements = []

    title_style = styles["Heading1"]

    heading_style = styles["Heading2"]

    normal = styles["BodyText"]

    elements.append(
        Paragraph("FinSight Financial Report", title_style)
    )

    elements.append(
        Paragraph("<br/>", normal)
    )

    elements.append(
        Paragraph(
            f"<b>Name:</b> {current_user.fullname}",
            normal
        )
    )

    elements.append(
        Paragraph(
            f"<b>Email:</b> {current_user.email}",
            normal
        )
    )

    elements.append(
        Paragraph("<br/>", normal)
    )

    elements.append(
        Paragraph("Executive Summary", heading_style)
    )

    summary = [

        ["Metric", "Value"],

        ["Total Income", f"Rs {report['total_income']:,.2f}"],

        ["Total Expense", f"Rs {report['total_expense']:,.2f}"],

        ["Total Savings", f"Rs {report['total_savings']:,.2f}"],

        ["Investment Value", f"Rs {report['total_investment']:,.2f}"],

        ["Net Worth", f"RS {report['net_worth']:,.2f}"],

        ["Savings Rate", f"{report['savings_rate']} %"],

        ["Health Score", f"{report['health_score']}/100"]

    ]

    table = Table(summary, colWidths=[220, 220])

    table.setStyle(TableStyle([

        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2563EB")),

        ("TEXTCOLOR", (0,0), (-1,0), colors.white),

        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),

        ("BACKGROUND", (0,1), (0,-1), colors.whitesmoke),

        ("BOTTOMPADDING", (0,0), (-1,-1), 8),

        ("TOPPADDING", (0,0), (-1,-1), 8),

        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),

    ]))

    elements.append(table)

    elements.append(
        Paragraph("<br/><br/>", normal)
    )

    # -------------------------
    # Income Details
    # -------------------------

    elements.append(
        Paragraph("Income Details", heading_style)
    )

    income_data = [
        ["Source", "Amount", "Date", "Description"]
    ]

    for income in report["incomes"]:
        income_data.append([
            income.source,
            f"Rs. {income.amount:,.2f}",
            income.income_date.strftime("%d-%m-%Y"),
            income.description or "-"
        ])

    if len(income_data) == 1:
        income_data.append(["No Income Records", "-", "-", "-"])

    income_table = Table(
        income_data,
        colWidths=[120, 100, 100, 170]
    )

    income_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BACKGROUND", (0,1), (-1,-1), colors.beige),
    ]))

    elements.append(income_table)

    elements.append(
        Paragraph("<br/><br/>", normal)
    )

    # -------------------------
    # Expense Details
    # -------------------------

    elements.append(
        Paragraph("Expense Details", heading_style)
    )

    expense_data = [
        ["Category", "Amount", "Date", "Payment", "Description"]
    ]

    for expense in report["expenses"]:
        expense_data.append([
            expense.category,
            f"Rs. {expense.amount:,.2f}",
            expense.expense_date.strftime("%d-%m-%Y"),
            expense.payment_method,
            expense.description or "-"
        ])

    if len(expense_data) == 1:
        expense_data.append(["No Expense Records", "-", "-", "-", "-"])

    expense_table = Table(
        expense_data,
        colWidths=[90, 80, 90, 100, 160]
    )

    expense_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#DC2626")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BACKGROUND", (0,1), (-1,-1), colors.whitesmoke),
    ]))

    elements.append(expense_table)

    elements.append(
        Paragraph("<br/><br/>", normal)
    )

    # -------------------------
    # Budget Details
    # -------------------------

    elements.append(
        Paragraph("Budget Details", heading_style)
    )

    budget_data = [
        ["Month", "Year", "Budget Amount"]
    ]

    for budget in report["budgets"]:
        budget_data.append([
            budget.budget_month,
            str(budget.budget_year),
            f"Rs. {budget.budget_amount:,.2f}"
        ])

    if len(budget_data) == 1:
        budget_data.append(["-", "-", "No Budget Records"])

    budget_table = Table(
        budget_data,
        colWidths=[180, 120, 180]
    )

    budget_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#16A34A")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BACKGROUND", (0,1), (-1,-1), colors.beige),
    ]))

    elements.append(budget_table)

    elements.append(
        Paragraph("<br/><br/>", normal)
    )

    # -------------------------
    # Goals Progress
    # -------------------------

    elements.append(
        Paragraph("Financial Goals", heading_style)
    )

    goal_data = [
        ["Goal", "Target", "Saved", "Progress", "Target Date"]
    ]

    for goal in report["goals"]:

        progress = 0

        if goal.target_amount > 0:
            progress = round(
                (goal.saved_amount / goal.target_amount) * 100,
                1
            )

        goal_data.append([
            goal.goal_name,
            f"Rs. {goal.target_amount:,.2f}",
            f"Rs. {goal.saved_amount:,.2f}",
            f"{progress}%",
            goal.target_date.strftime("%d-%m-%Y")
        ])

    if len(goal_data) == 1:
        goal_data.append(["No Goals", "-", "-", "-", "-"])

    goal_table = Table(
        goal_data,
        colWidths=[130, 90, 90, 70, 120]
    )

    goal_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#9333EA")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BACKGROUND", (0,1), (-1,-1), colors.whitesmoke),
    ]))

    elements.append(goal_table)

    elements.append(
        Paragraph("<br/><br/>", normal)
    )

    # -------------------------
    # Investment Details
    # -------------------------

    elements.append(
        Paragraph("Investment Details", heading_style)
    )

    investment_data = [
        ["Investment", "Type", "Quantity", "Current Value"]
    ]

    for investment in report["investments"]:

        current_value = investment.quantity * investment.current_price

        investment_data.append([
            investment.investment_name,
            investment.investment_type,
            str(investment.quantity),
            f"Rs. {current_value:,.2f}"
        ])

    if len(investment_data) == 1:
        investment_data.append(["No Investments", "-", "-", "-"])

    investment_table = Table(
        investment_data,
        colWidths=[160, 120, 80, 140]
    )

    investment_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F59E0B")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BACKGROUND", (0,1), (-1,-1), colors.beige),
    ]))

    elements.append(investment_table)

    elements.append(
        Paragraph("<br/><br/>", normal)
    )

    # -------------------------
    # Recent Notifications
    # -------------------------

    elements.append(
        Paragraph("Recent Notifications", heading_style)
    )

    notification_data = [
        ["Title", "Category", "Date"]
    ]

    for notification in report["notifications"][:10]:

        notification_data.append([
            notification.title,
            notification.category,
            notification.created_at.strftime("%d-%m-%Y")
        ])

    if len(notification_data) == 1:
        notification_data.append(["No Notifications", "-", "-"])

    notification_table = Table(
        notification_data,
        colWidths=[240, 120, 120]
    )

    notification_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BACKGROUND", (0,1), (-1,-1), colors.whitesmoke),
    ]))

    elements.append(notification_table)

    elements.append(
        Paragraph("<br/><br/>", normal)
    )

    from datetime import datetime

    elements.append(
        Paragraph(
            f"<b>Report Generated On:</b> {datetime.now().strftime('%d-%m-%Y %I:%M %p')}",
            normal
        )
    )

    elements.append(
        Paragraph(
            "Generated by FinSight - Personal Finance Management System",
            normal
        )
    )
    doc.build(elements)

    buffer.seek(0)

    # Return the PDF as a download
    return send_file(
        buffer,
        as_attachment=True,
        download_name="FinSight_Financial_Report.pdf",
        mimetype="application/pdf"
    )

# -------------------------
# export excel
# ------------------------
@app.route("/reports/export/excel")
@login_required
def export_excel():

    report = get_report_data(current_user.user_id)

    wb = Workbook()

    ws = wb.active

    ws.title = "Financial Report"

    title_font = Font(bold=True, size=16)

    heading_font = Font(bold=True)

    fill = PatternFill(
        start_color="2563EB",
        end_color="2563EB",
        fill_type="solid"
    )

    ws["A1"] = "FinSight Financial Report"
    ws["A1"].font = title_font

    ws["A3"] = "Name"
    ws["B3"] = current_user.fullname

    ws["A4"] = "Email"
    ws["B4"] = current_user.email

    row = 6

    ws.cell(row=row, column=1).value = "Executive Summary"
    ws.cell(row=row, column=1).font = heading_font

    row += 2

    summary = [

        ("Total Income", report["total_income"]),

        ("Total Expense", report["total_expense"]),

        ("Total Savings", report["total_savings"]),

        ("Investment Value", report["total_investment"]),

        ("Net Worth", report["net_worth"]),

        ("Savings Rate", f"{report['savings_rate']}%"),

        ("Health Score", report["health_score"])

    ]

    for metric, value in summary:

        ws.cell(row=row, column=1).value = metric

        ws.cell(row=row, column=2).value = value

        row += 1

    income_ws = wb.create_sheet("Income")

    income_ws.append([
        "Source",
        "Amount",
        "Date",
        "Description"
    ])

    for cell in income_ws[1]:
        cell.font = heading_font
        cell.fill = fill

    for income in report["incomes"]:

        income_ws.append([
            income.source,
            income.amount,
            income.income_date.strftime("%d-%m-%Y"),
            income.description
        ])

    expense_ws = wb.create_sheet("Expenses")

    expense_ws.append([
        "Category",
        "Amount",
        "Date",
        "Payment Method",
        "Description"
    ])

    for cell in expense_ws[1]:
        cell.font = heading_font
        cell.fill = fill

    for expense in report["expenses"]:

        expense_ws.append([
            expense.category,
            expense.amount,
            expense.expense_date.strftime("%d-%m-%Y"),
            expense.payment_method,
            expense.description
        ])

    budget_ws = wb.create_sheet("Budgets")

    budget_ws.append([
        "Month",
        "Year",
        "Budget Amount"
    ])

    for cell in budget_ws[1]:
        cell.font = heading_font
        cell.fill = fill

    for budget in report["budgets"]:

        budget_ws.append([
            budget.budget_month,
            budget.budget_year,
            budget.budget_amount
        ])

    goal_ws = wb.create_sheet("Goals")

    goal_ws.append([
        "Goal",
        "Target",
        "Saved",
        "Target Date"
    ])

    for cell in goal_ws[1]:
        cell.font = heading_font
        cell.fill = fill

    for goal in report["goals"]:

        goal_ws.append([
            goal.goal_name,
            goal.target_amount,
            goal.saved_amount,
            goal.target_date.strftime("%d-%m-%Y")
        ])

    investment_ws = wb.create_sheet("Investments")

    investment_ws.append([
        "Investment",
        "Type",
        "Quantity",
        "Buy Price",
        "Current Price"
    ])

    for cell in investment_ws[1]:
        cell.font = heading_font
        cell.fill = fill

    for investment in report["investments"]:

        investment_ws.append([
            investment.investment_name,
            investment.investment_type,
            investment.quantity,
            investment.buy_price,
            investment.current_price
        ])

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="FinSight_Financial_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# -------------------------
# SETTINGS
# -------------------------

@app.route("/settings")
@login_required
def settings():

    settings = UserSettings.query.filter_by(
        user_id=current_user.user_id
    ).first()

    if settings is None:
        settings = UserSettings(user_id=current_user.user_id)
        db.session.add(settings)
        db.session.commit()

    return render_template(
        "settings.html",
        settings=settings
    )
# -------------------------
# UPDATE SETTINGS
# -------------------------

@app.route("/settings/update", methods=["POST"])
@login_required
def update_settings():

    settings = UserSettings.query.filter_by(
        user_id=current_user.user_id
    ).first()

    selected_currency = request.form.get("currency")

    try:

        if selected_currency == "INR":
            exchange_rate = 1.0

        else:

            url = f"https://api.frankfurter.app/latest?from=INR&to={selected_currency}"

            response = requests.get(url, timeout=10)

            data = response.json()

            exchange_rate = data["rates"][selected_currency]

        settings.currency = selected_currency
        settings.exchange_rate = exchange_rate
        settings.last_updated = datetime.now()

    except Exception:

        flash(
            "Unable to fetch latest exchange rates. Please connect to the internet and try again.",
            "danger"
        )

        return redirect(url_for("settings"))

    settings.theme = request.form.get("theme")

    settings.budget_alert = "budget_alert" in request.form
    settings.goal_reminder = "goal_reminder" in request.form
    settings.investment_update = "investment_update" in request.form
    settings.monthly_report = "monthly_report" in request.form

    db.session.commit()

    flash("Settings updated successfully!", "success")

    return redirect(url_for("settings"))

@app.route("/delete-account")
@login_required
def delete_account():

    return "Delete Account feature coming soon!"

# -------------------------
# FORGOT PASSWORD
# -------------------------

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():

    if request.method == "POST":

        email = request.form["email"]
        new_password = request.form["new_password"]
        confirm_password = request.form["confirm_password"]

        if new_password != confirm_password:

            flash("Passwords do not match!", "danger")
            return redirect("/forgot-password")

        user = User.query.filter_by(email=email).first()

        if not user:

            flash("Email not found!", "danger")
            return redirect("/forgot-password")

        user.password = generate_password_hash(new_password)

        db.session.commit()

        flash("Password reset successfully. Please login.", "success")

        return redirect("/login")

    return render_template("forgot_password.html")

# -------------------------
# PROFILE
# -------------------------

@app.route("/profile")
@login_required
def profile():

    return render_template(
        "profile.html",
        user=current_user
    )


@app.route("/update-profile", methods=["POST"])
@login_required
def update_profile():

    current_user.fullname = request.form["fullname"]
    current_user.mobile = request.form["mobile"]
    current_user.occupation = request.form["occupation"]

    income = request.form["monthly_income"]
    current_user.monthly_income = float(income) if income else None

    current_user.income_source = request.form["income_source"]
    current_user.risk_level = request.form["risk_level"]
    current_user.financial_goal = request.form["financial_goal"]

    # ---------- Profile Image Upload ----------

    file = request.files.get("profile_image")

    if file and file.filename != "":

        filename = secure_filename(file.filename)

        upload_folder = os.path.join(app.static_folder, "uploads")

        os.makedirs(upload_folder, exist_ok=True)

        file.save(os.path.join(upload_folder, filename))

        current_user.profile_image = filename

    # ------------------------------------------

    db.session.commit()

    flash("Profile updated successfully!", "success")

    return redirect("/profile")
# -------------------------
# REGISTER
# -------------------------

@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":

        fullname = request.form["fullname"]
        email = request.form["email"]
        mobile = request.form["mobile"]
        password = request.form["password"]

        user = User.query.filter_by(email=email).first()

        if user:
            flash("Email already exists!", "danger")
            return redirect("/register")

        password = generate_password_hash(password)

        new_user = User(
            fullname=fullname,
            email=email,
            mobile=mobile,
            password=password,
            is_verified=False
        )

        db.session.add(new_user)
        db.session.commit()

        token = serializer.dumps(
            new_user.email,
            salt="email-verification"
        )

        verification_url = url_for(
            "verify_email",
            token=token,
            _external=True
        )

        message = Message(
            subject="Verify Your Email - FinSight",
            recipients=[new_user.email]
        )

        message.html = f"""
        <h2>Welcome to FinSight!</h2>

        <p>Hello {new_user.fullname},</p>

        <p>Please verify your email address to activate your account.</p>

        <a href="{verification_url}"
        style="
            background-color:#198754;
            color:white;
            padding:12px 20px;
            text-decoration:none;
            border-radius:5px;
            display:inline-block;
        ">
            Verify Email
        </a>

        <p>This verification link expires in 1 hour.</p>

        <p>FinSight Team</p>
        """

        mail.send(message)

        flash(
            "Registration successful. Please verify your email before login.",
            "success"
        )

        return redirect("/login")

    return render_template("register.html")

# -------------------------
# email verification
# -------------------------

@app.route("/verify-email/<token>")
def verify_email(token):

    try:

        email = serializer.loads(
            token,
            salt="email-verification",
            max_age=3600
        )

    except Exception:

        flash(
            "Verification link is invalid or expired.",
            "danger"
        )

        return redirect(url_for("login"))

    user = User.query.filter_by(
        email=email
    ).first()

    if not user:

        flash("User not found.", "danger")

        return redirect(url_for("register"))

    user.is_verified = True

    db.session.commit()

    flash(
        "Email verified successfully! You can now login.",
        "success"
    )

    return redirect(url_for("login"))


# -------------------------
# LOGIN
# -------------------------

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        email = request.form["email"]
        password = request.form["password"]

        user = User.query.filter_by(email=email).first()

        if user and check_password_hash(user.password, password):

            if not user.is_verified:

                flash(
                    "Please verify your email before logging in.",
                    "warning"
                )

                return redirect("/login")

            login_user(user)

            generate_notifications(user.user_id)

            flash("Login Successful!", "success")

            return redirect("/dashboard")

        flash("Invalid Email or Password!", "danger")

    return render_template("login.html")

# -------------------------
# LOGOUT
# -------------------------

@app.route("/logout")
@login_required
def logout():

    logout_user()

    flash("Logged out successfully!", "success")

    return redirect("/login")

# -------------------------
# CHANGE PASSWORD
# -------------------------
@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():

    if request.method == "POST":

        current_password = request.form["current_password"]
        new_password = request.form["new_password"]
        confirm_password = request.form["confirm_password"]

        if not check_password_hash(
            current_user.password,
            current_password
        ):

            flash(
                "Current password is incorrect.",
                "danger"
            )

            return redirect(
                url_for("change_password")
            )
        
        if check_password_hash(
            current_user.password,
            new_password
        ):

            flash(
                "New password cannot be the same as the current password.",
                "warning"
            )

            return redirect(
                url_for("change_password")
            )

        if new_password != confirm_password:

            flash(
                "New passwords do not match.",
                "danger"
            )

            return redirect(
                url_for("change_password")
            )

        current_user.password = generate_password_hash(
            new_password
        )



        db.session.commit()

        flash(
            "Password changed successfully!",
            "success"
        )

        return redirect(
            url_for("dashboard")
        )

    return render_template(
        "change_password.html"
    )


# ==========================================================
# SPENDING ANALYSIS
# ==========================================================

@app.route("/analytics/spending")
@login_required
def spending_analysis():

    expenses = Expense.query.filter_by(
        user_id=current_user.user_id
    ).all()

    # ------------------------------------------------------
    # CATEGORY WISE EXPENSES
    # ------------------------------------------------------

    category_data = {}

    for expense in expenses:

        category = expense.category

        category_data[category] = (
            category_data.get(category, 0)
            + expense.amount
        )

    labels = list(category_data.keys())

    values = list(category_data.values())

    # ------------------------------------------------------
    # TOTAL EXPENSE
    # ------------------------------------------------------

    total_expense = sum(values)

    # ------------------------------------------------------
    # HIGHEST CATEGORY
    # ------------------------------------------------------

    if category_data:

        highest_category = max(
            category_data,
            key=category_data.get
        )

        highest_amount = category_data[
            highest_category
        ]

    else:

        highest_category = "N/A"

        highest_amount = 0

    # ------------------------------------------------------
    # HIGHEST CATEGORY PERCENTAGE
    # ------------------------------------------------------

    highest_percentage = 0

    if total_expense > 0 and highest_amount > 0:
        highest_percentage = (highest_amount / total_expense) * 100

    # ------------------------------------------------------
    # TOTAL CATEGORIES
    # ------------------------------------------------------

    total_categories = len(category_data)

    # ------------------------------------------------------
    # LOWEST CATEGORY
    # ------------------------------------------------------

    if category_data:

        lowest_category = min(
            category_data,
            key=category_data.get
        )

        lowest_amount = category_data[
            lowest_category
        ]

    else:

        lowest_category = "N/A"

        lowest_amount = 0

    # ------------------------------------------------------
    # AVERAGE DAILY EXPENSE
    # ------------------------------------------------------

    if expenses:

        unique_days = len(

            set(

                expense.expense_date

                for expense in expenses

            )

        )

        average_daily = (

            total_expense / unique_days

            if unique_days > 0

            else 0

        )

    else:

        average_daily = 0

    # ------------------------------------------------------
    # TOP CATEGORIES
    # ------------------------------------------------------

    top_categories = sorted(

        category_data.items(),

        key=lambda x: x[1],

        reverse=True

    )

    # ------------------------------------------------------
    # RENDER
    # ------------------------------------------------------

    total_expense = convert_amount(total_expense)
    highest_amount = convert_amount(highest_amount)
    lowest_amount = convert_amount(lowest_amount)
    average_daily = convert_amount(average_daily)

    converted_top_categories = []

    for category, amount in top_categories:
        converted_top_categories.append(
            (category, convert_amount(amount))
        )

    top_categories = converted_top_categories

    values = convert_list(values)

    return render_template(

        "spending_analysis.html",

        labels=labels,

        values=values,

        total_expense=total_expense,

        highest_category=highest_category,

        highest_amount=highest_amount,

        highest_percentage=highest_percentage,

        total_categories=total_categories,

        lowest_category=lowest_category,

        lowest_amount=lowest_amount,

        average_daily=average_daily,

        top_categories=top_categories,

        expense_categories=category_data,

        currency_symbol=get_currency_symbol()

    )

# ==========================================================
# END SPENDING ANALYSIS
# ==========================================================


# ==========================================================
# BUDGET RECOMMENDATIONS
# ==========================================================

@app.route("/analytics/recommendations")
@login_required
def budget_recommendations():

    expenses = Expense.query.filter_by(
        user_id=current_user.user_id
    ).all()

    budgets = Budget.query.filter_by(
        user_id=current_user.user_id
    ).all()

    incomes = Income.query.filter_by(
        user_id=current_user.user_id
    ).all()

    total_income = sum(
        income.amount
        for income in incomes
    )

    total_expense = sum(
        expense.amount
        for expense in expenses
    )

    total_budget = sum(
        budget.budget_amount
        for budget in budgets
    )

    recommendations = []

    # ------------------------------------------------------
    # Budget Check
    # ------------------------------------------------------

    if total_budget > 0:

        utilization = (
            total_expense / total_budget
        ) * 100

        if utilization > 100:

            recommendations.append({

                "title":"Budget Exceeded",

                "message":"Reduce unnecessary expenses to stay within your budget.",

                "type":"danger"

            })

        elif utilization > 80:

            recommendations.append({

                "title":"Budget Warning",

                "message":"You have used more than 80% of your budget.",

                "type":"warning"

            })

        else:

            recommendations.append({

                "title":"Excellent Budget Control",

                "message":"Your spending is well within your budget.",

                "type":"success"

            })

    # ------------------------------------------------------
    # Savings
    # ------------------------------------------------------

    savings = total_income - total_expense

    if savings <= 0:

        recommendations.append({

            "title":"Increase Savings",

            "message":"Try reducing discretionary expenses.",

            "type":"danger"

        })

    elif savings < (0.20 * total_income):

        recommendations.append({

            "title":"Improve Savings",

            "message":"Aim to save at least 20% of your income.",

            "type":"warning"

        })

    else:

        recommendations.append({

            "title":"Healthy Savings",

            "message":"Great job maintaining a healthy savings rate.",

            "type":"success"

        })

    # ------------------------------------------------------
    # Highest Expense Category
    # ------------------------------------------------------

    category_totals = {}

    for expense in expenses:

        category_totals[expense.category] = (

            category_totals.get(

                expense.category,

                0

            )

            + expense.amount

        )

    highest_category = None

    highest_amount = 0

    if category_totals:

        highest_category = max(

            category_totals,

            key=category_totals.get

        )

        highest_amount = category_totals[

            highest_category

        ]

        recommendations.append({

            "title":"Highest Spending",

            "message":

            f"You spend the most on "

            f"{highest_category} "

            f"(₹ {highest_amount:.2f}).",

            "type":"info"

        })

    # ------------------------------------------------------
    # Investment Recommendation
    # ------------------------------------------------------

    investment_count = Investment.query.filter_by(
        user_id=current_user.user_id
    ).count()

    if investment_count == 0:

        recommendations.append({

            "title": "Start Investing",

            "message": "You have not added any investments. Investing can help grow your wealth over time.",

            "type": "info"

        })

    else:

        recommendations.append({

            "title": "Investment Portfolio",

            "message": f"You currently have {investment_count} investment(s). Continue monitoring their performance regularly.",

            "type": "success"

        })


    # ------------------------------------------------------
    # Financial Goal Recommendation
    # ------------------------------------------------------

    goal_count = Goal.query.filter_by(
        user_id=current_user.user_id
    ).count()

    if goal_count == 0:

        recommendations.append({

            "title": "Create a Financial Goal",

            "message": "Set a financial goal to better plan and track your future savings.",

            "type": "warning"

        })

    else:

        recommendations.append({

            "title": "Financial Goals",

            "message": f"You have {goal_count} financial goal(s). Keep tracking your progress.",

            "type": "success"

        })

    if total_budget > 0:
        utilization = (total_expense / total_budget) * 100
    else:
        utilization = 0

    tips = []

    # Savings Tips
    if savings < total_income * 0.20:
        tips.append({
            "icon": "fa-piggy-bank",
            "title": "Increase Savings",
            "message": "Try saving at least 20% of your monthly income."
        })

    # Budget Tips
    if utilization >= 80:
        tips.append({
            "icon": "fa-chart-pie",
            "title": "Review Budget",
            "message": "You are close to your budget limit. Monitor your spending."
        })

    # Highest Spending Category
    if highest_category:
        tips.append({
            "icon": "fa-credit-card",
            "title": "Reduce Spending",
            "message": f"Review your '{highest_category}' expenses to reduce costs."
        })

    # Investment Tips
    if investment_count == 0:
        tips.append({
            "icon": "fa-chart-line",
            "title": "Start Investing",
            "message": "Consider starting a SIP or mutual fund investment."
        })

    # Emergency Fund
    if savings > 0:
        tips.append({
            "icon": "fa-shield-heart",
            "title": "Emergency Fund",
            "message": "Maintain 3–6 months of expenses as an emergency fund."
        })

    # Goal Tips
    if goal_count == 0:
        tips.append({
            "icon": "fa-bullseye",
            "title": "Create Goals",
            "message": "Set financial goals to track your future plans."
        })

    total_income = convert_amount(total_income)
    total_expense = convert_amount(total_expense)
    total_budget = convert_amount(total_budget)
    savings = convert_amount(savings)

    return render_template(

        "budget_recommendations.html",

        tips=tips,

        utilization=utilization,

        total_income=total_income,

        total_expense=total_expense,

        total_budget=total_budget,

        savings=savings,

        recommendations=recommendations,

        currency_symbol=get_currency_symbol()

    )

# ==========================================================
# END BUDGET RECOMMENDATIONS
# ==========================================================


# ==========================================================
# HEALTH SCORE
# ==========================================================

@app.route("/analytics/health-score")
@login_required
def health_score():

    incomes = Income.query.filter_by(
        user_id=current_user.user_id
    ).all()

    expenses = Expense.query.filter_by(
        user_id=current_user.user_id
    ).all()

    budgets = Budget.query.filter_by(
        user_id=current_user.user_id
    ).all()

    investments = Investment.query.filter_by(
        user_id=current_user.user_id
    ).all()

    goals = Goal.query.filter_by(
        user_id=current_user.user_id
    ).all()

    # ------------------------------------------------------
    # TOTALS
    # ------------------------------------------------------

    total_income = sum(
        income.amount
        for income in incomes
    )

    total_expense = sum(
        expense.amount
        for expense in expenses
    )

    total_budget = sum(
        budget.budget_amount
        for budget in budgets
    )

    total_investment = sum(
        investment.quantity *
        investment.current_price
        for investment in investments
    )

    total_goals = len(goals)

    completed_goals = sum(
        1
        for goal in goals
        if goal.saved_amount >= goal.target_amount
    )

    # ------------------------------------------------------
    # HEALTH SCORE CALCULATION
    # ------------------------------------------------------

    health_score = 100

    if total_income > 0:

        expense_ratio = (
            total_expense /
            total_income
        ) * 100

        if expense_ratio > 90:

            health_score -= 30

        elif expense_ratio > 75:

            health_score -= 20

        elif expense_ratio > 60:

            health_score -= 10

    if total_budget > 0 and total_expense > total_budget:

        health_score -= 15

    if total_investment == 0:

        health_score -= 15

    if total_goals == 0:

        health_score -= 10

    health_score = max(
        health_score,
        0
    )

    # ------------------------------------------------------
    # HEALTH STATUS
    # ------------------------------------------------------

    if health_score >= 90:

        health_status = "Excellent"

    elif health_score >= 75:

        health_status = "Good"

    elif health_score >= 60:

        health_status = "Average"

    elif health_score >= 40:

        health_status = "Needs Improvement"

    else:

        health_status = "Poor"

    # ------------------------------------------------------
    # Financial Health Tips
    # ------------------------------------------------------

    health_tips = []

    if total_expense > total_income:

        health_tips.append(
            "Reduce your monthly expenses to avoid overspending."
        )

    if total_budget > 0 and total_expense > total_budget:

        health_tips.append(
            "Your expenses have exceeded your budget."
        )

    if total_income > 0 and (total_income - total_expense) < (0.20 * total_income):

        health_tips.append(
            "Try to save at least 20% of your monthly income."
        )

    investment_count = Investment.query.filter_by(
        user_id=current_user.user_id
    ).count()

    if investment_count == 0:

        health_tips.append(
            "Consider starting your investment journey."
        )

    goal_count = Goal.query.filter_by(
        user_id=current_user.user_id
    ).count()

    if goal_count == 0:

        health_tips.append(
            "Create financial goals to stay focused."
        )

    if not health_tips:

        health_tips.append(
            "Excellent! Your financial habits look healthy."
        )

    # ==========================================================
    # PORTFOLIO INSIGHTS
    # ==========================================================

    investments = Investment.query.filter_by(
        user_id=current_user.user_id
    ).all()

    portfolio_value = 0
    best_investment = None
    worst_investment = None

    if investments:

        portfolio_value = sum(
            investment.quantity * investment.current_price
            for investment in investments
        )

        

        unique_assets = len(
            set(
                investment.symbol
                for investment in investments
            )
        )

        if unique_assets >= 5:

            diversification = "High"

            risk_level = "Low"

        elif unique_assets >= 3:

            diversification = "Moderate"

            risk_level = "Medium"

        else:

            diversification = "Low"

            risk_level = "High"

    else:

        diversification = "No Investments"

        risk_level = "Not Available"
    # ------------------------------------------------------
    # RENDER PAGE
    # ------------------------------------------------------
    portfolio_value = convert_amount(portfolio_value)

    return render_template(

        "health_score.html",

        health_score=health_score,

        health_status=health_status,

        total_income=total_income,

        total_expense=total_expense,

        total_budget=total_budget,

        total_investment=total_investment,

        total_goals=total_goals,

        completed_goals=completed_goals,

        health_tips=health_tips,

        portfolio_value=portfolio_value,

        best_investment=None,

        worst_investment=None,

        diversification=diversification,

        risk_level=risk_level,

        currency_symbol=get_currency_symbol()

    )

# ==========================================================
# END HEALTH SCORE
# ==========================================================

# ==========================================================
# TRENDS & PREDICTIONS
# ==========================================================

@app.route("/analytics/trends")
@login_required
def trends_predictions():

    incomes = Income.query.filter_by(
        user_id=current_user.user_id
    ).all()

    expenses = Expense.query.filter_by(
        user_id=current_user.user_id
    ).all()

    months = [
        "Jan","Feb","Mar","Apr",
        "May","Jun","Jul","Aug",
        "Sep","Oct","Nov","Dec"
    ]

    income_data = [0] * 12
    expense_data = [0] * 12
    savings_data = [0] * 12

    for income in incomes:

        month = income.income_date.month - 1

        income_data[month] += income.amount

    for expense in expenses:

        month = expense.expense_date.month - 1

        expense_data[month] += expense.amount

    for i in range(12):

        savings_data[i] = (
            income_data[i]
            - expense_data[i]
        )

    total_income = sum(income_data)

    total_expense = sum(expense_data)

    total_savings = (
        total_income
        - total_expense
    )

    predicted_savings = (
        total_savings * 1.10
    )

    # ==========================================================
    # MONTHLY SAVINGS RATE
    # ==========================================================

    if total_income > 0:

        savings_rate = (
            total_savings / total_income
        ) * 100

    else:

        savings_rate = 0

    if savings_rate >= 20:

        savings_status = "Excellent"

    elif savings_rate >= 10:

        savings_status = "Good"

    elif savings_rate > 0:

        savings_status = "Needs Improvement"

    else:

        savings_status = "Poor"

    # ==========================================================
    # FINANCIAL SUMMARY
    # ==========================================================

    investments = Investment.query.filter_by(
        user_id=current_user.user_id
    ).all()

    total_investment = 0

    for investment in investments:
        total_investment += (
            investment.quantity * investment.current_price
        )

    net_worth = total_investment  + total_savings

    if total_income > 0:

        expense_ratio = (
            total_expense / total_income
        ) * 100

    else:

        expense_ratio = 0

    if expense_ratio <= 50:

        expense_status = "Excellent"

    elif expense_ratio <= 80:

        expense_status = "Good"

    else:

        expense_status = "High Spending"

    # ==========================================================
    # RECENT FINANCIAL ACTIVITIES
    # ==========================================================

    activities = []

    for income in incomes:

        activities.append({

            "date": income.income_date,

            "type": "Income",

            "description": f"Added Income ₹{income.amount:.2f}"

        })

    for expense in expenses:

        activities.append({

            "date": expense.expense_date,

            "type": "Expense",

            "description": f"Added Expense ₹{expense.amount:.2f}"

        })

    investments = Investment.query.filter_by(
        user_id=current_user.user_id
    ).all()

    for investment in investments:

        activities.append({

            "date": investment.purchase_date,

            "type": "Investment",

            "description": f"Purchased {investment.symbol}"

        })

    goals = Goal.query.filter_by(
        user_id=current_user.user_id
    ).all()

    for goal in goals:

        activities.append({

            "date": goal.created_at,

            "type": "Goal",

            "description": f"Created Goal: {goal.goal_name}"

        })

    from datetime import datetime

    activities.sort(
        key=lambda activity: (
            datetime.combine(activity["date"], datetime.min.time())
            if not isinstance(activity["date"], datetime)
            else activity["date"]
        ),
        reverse=True
    )

    activities = activities[:5]

    months_with_income = max(1, len([x for x in income_data if x > 0]))
    months_with_expense = max(1, len([x for x in expense_data if x > 0]))
    months_with_savings = max(1, len([x for x in savings_data if x != 0]))

    avg_income = total_income / months_with_income
    avg_expense = total_expense / months_with_expense
    avg_savings = total_savings / months_with_savings

    non_zero = [x for x in expense_data if x > 0]

    if non_zero:
        predicted_expense = sum(non_zero[-3:]) / min(3, len(non_zero))
    else:
        predicted_expense = 0

    avg_income = convert_amount(avg_income)
    avg_expense = convert_amount(avg_expense)
    avg_savings = convert_amount(avg_savings)
    predicted_expense = convert_amount(predicted_expense)

    income_data = convert_list(income_data)
    expense_data = convert_list(expense_data)
    savings_data = convert_list(savings_data)

    return render_template(

        "trends_predictions.html",

        months=months,

        income_data=income_data,

        expense_data=expense_data,

        savings_data=savings_data,

        avg_income=avg_income,

        avg_expense=avg_expense,

        avg_savings=avg_savings,

        predicted_expense=predicted_expense,

        total_income=total_income,

        total_expense=total_expense,

        total_savings=total_savings,

        savings_rate=savings_rate,

        savings_status=savings_status,

        expense_ratio=expense_ratio,

        activities=activities,

        currency_symbol=get_currency_symbol()

    )

# ==========================================================
# END TRENDS & PREDICTIONS
# ==========================================================

# ==========================================================
# NOTIFICATIONS
# ==========================================================
@app.route("/notifications")
@login_required
def notifications():

    notifications = Notification.query.filter_by(
        user_id=current_user.user_id
    ).order_by(
        Notification.created_at.desc()
    ).all()

    unread_count = Notification.query.filter_by(
        user_id=current_user.user_id,
        is_read=False
    ).count()

    return render_template(
        "notifications.html",
        notifications=notifications,
        unread_count=unread_count
    )

# ==========================================================
# END NOTIFICATIONS
# ==========================================================
@app.route("/notification/read/<int:notification_id>")
@login_required
def mark_notification_read(notification_id):

    notification = Notification.query.filter_by(
        notification_id=notification_id,
        user_id=current_user.user_id
    ).first_or_404()

    notification.is_read = True

    db.session.commit()

    return redirect(url_for("notifications"))


@app.route("/notifications/read-all")
@login_required
def mark_all_notifications_read():

    Notification.query.filter_by(
        user_id=current_user.user_id,
        is_read=False
    ).update({"is_read": True})

    db.session.commit()

    flash("All notifications marked as read.", "success")

    return redirect(url_for("notifications"))


@app.route("/notification/delete/<int:notification_id>")
@login_required
def delete_notification(notification_id):

    notification = Notification.query.filter_by(
        notification_id=notification_id,
        user_id=current_user.user_id
    ).first_or_404()

    db.session.delete(notification)

    db.session.commit()

    flash("Notification deleted.", "success")

    return redirect(url_for("notifications"))


@app.route("/notifications/clear")
@login_required
def clear_notifications():

    Notification.query.filter_by(
        user_id=current_user.user_id
    ).delete()

    db.session.commit()

    flash("All notifications cleared.", "success")

    return redirect(url_for("notifications"))

# ==========================================================
# EDIT INCOME
# ==========================================================

@app.route("/edit-income/<int:income_id>", methods=["GET", "POST"])
@login_required
def edit_income(income_id):

    income = Income.query.filter_by(
        income_id=income_id,
        user_id=current_user.user_id
    ).first_or_404()

    if request.method == "POST":

        income.source = request.form["source"]

        income.amount = float(
            request.form["amount"]
        )

        income.income_date = request.form["income_date"]

        income.description = request.form["description"]

        db.session.commit()

        flash(
            "Income updated successfully!",
            "success"
        )

        return redirect(url_for("income"))
    income.converted_amount = convert_amount(income.amount)
    return render_template(
        "edit_income.html",
        income=income,
        currency_symbol=get_currency_symbol()
    )

# ==========================================================
# END EDIT INCOME
# ==========================================================

# ==========================================================
# DELETE INCOME
# ==========================================================

@app.route("/delete-income/<int:income_id>")
@login_required
def delete_income(income_id):

    income = Income.query.filter_by(
        income_id=income_id,
        user_id=current_user.user_id
    ).first_or_404()

    db.session.delete(income)

    db.session.commit()

    flash(
        "Income deleted successfully!",
        "success"
    )

    return redirect(
        url_for("income")
    )

# ==========================================================
# END DELETE INCOME
# ==========================================================

# ==========================================================
# EDIT BUDGET
# ==========================================================

@app.route("/edit-budget/<int:budget_id>", methods=["GET", "POST"])
@login_required
def edit_budget(budget_id):

    budget = Budget.query.filter_by(
        budget_id=budget_id,
        user_id=current_user.user_id
    ).first_or_404()

    if request.method == "POST":

        budget.budget_amount = float(
            request.form["budget_amount"]
        )

        budget.budget_year = request.form["budget_year"]

        db.session.commit()

        flash(
            "Budget updated successfully!",
            "success"
        )

        return redirect(url_for("budgets"))
    budget.converted_amount = convert_amount(budget.budget_amount)
    return render_template(
        "edit_budget.html",
        budget=budget,
        currency_symbol=get_currency_symbol()
    )

# ==========================================================
# END EDIT BUDGET
# ==========================================================

# ==========================================================
# DELETE BUDGET
# ==========================================================

@app.route("/delete-budget/<int:budget_id>")
@login_required
def delete_budget(budget_id):

    budget = Budget.query.filter_by(
        budget_id=budget_id,
        user_id=current_user.user_id
    ).first_or_404()

    db.session.delete(budget)

    db.session.commit()

    flash(
        "Budget deleted successfully!",
        "success"
    )

    return redirect(url_for("budgets"))

# ==========================================================
# END DELETE BUDGET
# ==========================================================
@app.context_processor
def inject_sidebar_data():
    if not current_user.is_authenticated:
        return {}

    from datetime import date

    today = date.today()

    monthly_income = db.session.query(
        db.func.sum(Income.amount)
    ).filter(
        Income.user_id == current_user.user_id,
        db.extract("month", Income.income_date) == today.month,
        db.extract("year", Income.income_date) == today.year
    ).scalar() or 0

    monthly_expense = db.session.query(
        db.func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.user_id,
        db.extract("month", Expense.expense_date) == today.month,
        db.extract("year", Expense.expense_date) == today.year
    ).scalar() or 0

    monthly_savings = monthly_income - monthly_expense

    return {
        "monthly_income": convert_amount(monthly_income),
        "monthly_expense": convert_amount(monthly_expense),
        "monthly_savings": convert_amount(monthly_savings),
        "currency_symbol": get_currency_symbol()
    }



@app.context_processor
def inject_theme():

    if current_user.is_authenticated:

        settings = UserSettings.query.filter_by(
            user_id=current_user.user_id
        ).first()

        if settings:
            return {"theme": settings.theme}

    return {"theme": "light"}
# -------------------------
# CREATE DATABASE TABLES
# -------------------------

with app.app_context():
    db.create_all()

print(app.url_map)

if __name__ == "__main__":
    app.run(debug=True)